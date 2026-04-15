import math
import logging
import pandas as pd
from qgis.core import QgsVectorLayer

logger = logging.getLogger(__name__)

# 논/답은 AMC3 = 79 고정
AMC3_FIXED_79 = {'논', '답'}


def _is_null(v) -> bool:
    """Python None, QVariant NULL, float NaN 모두 null로 처리."""
    if v is None:
        return True
    try:
        from qgis.PyQt.QtCore import QVariant
        if isinstance(v, QVariant) and v.isNull():
            return True
    except Exception:
        pass
    try:
        if isinstance(v, float) and math.isnan(v):
            return True
    except Exception:
        pass
    return False

HYDRO_TYPES = ['A', 'B', 'C', 'D']

# result1.xlsx 고정 토지이용 순서 (13개)
LAND_USE_ORDER = [
    '공급처리시설', '공업지', '공원', '광장',
    '논', '답', '밭', '상업지',
    '임야', '주거지', '주차장', '초지', '하천',
]


def layer_to_dataframe(layer: QgsVectorLayer) -> pd.DataFrame:
    """CN값_input 레이어를 DataFrame으로 변환."""
    records = []
    for feat in layer.getFeatures():
        area     = feat['유역면적']
        cn       = feat['cn값']
        hydro    = feat['토양군']
        records.append({
            '소유역명': feat['소유역명'],
            '토지이용': feat['토지이용'],
            '유역면적': float(area) if not _is_null(area) else 0.0,
            '토양군':   str(hydro).strip().upper() if not _is_null(hydro) else None,
            'cn값':     int(cn) if not _is_null(cn) else None,
        })
    return pd.DataFrame(records)


def _amc3(amc2: float, land_use: str) -> int:
    """AMC3 계산. 논/답 → 79 고정, 기타 → trunc(23×AMC2/(10+0.13×AMC2))."""
    if str(land_use).strip() in AMC3_FIXED_79:
        return 79
    if not amc2 or math.isnan(amc2) or amc2 <= 0:
        return 0
    return math.trunc(23.0 * amc2 / (10.0 + 0.13 * amc2))


def _calculate_watershed_cn(ws_name: str, ws_df: pd.DataFrame) -> tuple:
    """
    단일 유역(또는 합성 유역)의 CN값 계산.
    ws_df: DataFrame with columns [소유역명, 토지이용, 유역면적, 토양군, cn값]
    반환: (result1_entry: dict, result2_entry: dict)
    """
    # lu → {hydro: [total_area, cn_value]}
    lu_map = {}
    for _, row in ws_df.iterrows():
        lu = row['토지이용']
        ht = row['토양군']
        area = row['유역면적']
        cn = row['cn값']
        if lu not in lu_map:
            lu_map[lu] = {h: [0.0, None] for h in HYDRO_TYPES}
        if ht in HYDRO_TYPES:
            lu_map[lu][ht][0] += area
            if cn is not None:
                lu_map[lu][ht][1] = cn

    all_lus = sorted(lu_map.keys())
    ws_rows = []
    ws_total_area = 0.0
    ws_total_by_type = {h: 0.0 for h in HYDRO_TYPES}
    ws_cn_area_sum = 0.0
    ws_amc3_area_sum = 0.0

    for lu in all_lus:
        ht_data = lu_map.get(lu, {h: [0.0, None] for h in HYDRO_TYPES})
        row_total = sum(ht_data[h][0] for h in HYDRO_TYPES)
        cn_area_sum = sum(
            ht_data[h][0] * ht_data[h][1]
            for h in HYDRO_TYPES
            if ht_data[h][1] is not None
        )
        amc2 = cn_area_sum / row_total if row_total > 0 else 0.0
        amc3 = _amc3(amc2, lu) if row_total > 0 else 0

        ws_total_area += row_total
        ws_cn_area_sum += cn_area_sum
        ws_amc3_area_sum += amc3 * row_total
        for h in HYDRO_TYPES:
            ws_total_by_type[h] += ht_data[h][0]

        def _area(h, ht_data=ht_data):
            v = ht_data[h][0]
            return v if v > 0 else None

        ws_rows.append({
            'land_use': lu,
            'A_area': _area('A'), 'A_cn': ht_data['A'][1],
            'B_area': _area('B'), 'B_cn': ht_data['B'][1],
            'C_area': _area('C'), 'C_cn': ht_data['C'][1],
            'D_area': _area('D'), 'D_cn': ht_data['D'][1],
            'total_area': row_total if row_total > 0 else None,
            'amc2_cn': amc2 if row_total > 0 else None,
            'amc3_cn': amc3 if row_total > 0 else None,
        })

    ws_amc2 = ws_cn_area_sum / ws_total_area if ws_total_area > 0 else 0.0
    ws_amc3 = ws_amc3_area_sum / ws_total_area if ws_total_area > 0 else 0.0

    result1_entry = {
        'watershed': ws_name,
        'rows': ws_rows,
        'total_area': ws_total_area,
        'total_A': ws_total_by_type['A'],
        'total_B': ws_total_by_type['B'],
        'total_C': ws_total_by_type['C'],
        'total_D': ws_total_by_type['D'],
        'amc2_cn': ws_amc2,
        'amc3_cn': ws_amc3,
    }
    result2_entry = {
        'watershed': ws_name,
        'total_area': ws_total_area,
        'amc2_cn': ws_amc2,
        'amc3_cn': ws_amc3,
    }
    return result1_entry, result2_entry


def calculate_results(layer: QgsVectorLayer):
    """
    CN값_input 레이어에서 result1_data, result2_data 계산.

    result1_data: list of {
        'watershed', 'rows': [per-land-use row dicts],
        'total_area', 'total_A', 'total_B', 'total_C', 'total_D',
        'amc2_cn', 'amc3_cn'
    }
    result2_data: list of {'watershed', 'total_area', 'amc2_cn', 'amc3_cn'}
    """
    df = layer_to_dataframe(layer)
    if df.empty:
        raise ValueError("레이어에 데이터가 없습니다.")

    # CN값 NULL이면서 면적 있는 행 → 경고 목록
    null_mask = df['cn값'].isna() & (df['유역면적'] > 0)
    null_cn_rows = df[null_mask][['소유역명', '토지이용', '토양군']].values.tolist()
    if null_cn_rows:
        logger.warning(f"CN값 NULL 피처 {len(null_cn_rows)}건 — 해당 피처는 계산에서 제외됩니다.")

    # 소유역 순서 유지
    watersheds = list(dict.fromkeys(df['소유역명'].dropna().tolist()))
    result1_data = []
    result2_data = []

    for ws_name in watersheds:
        ws_df = df[df['소유역명'] == ws_name]
        r1_entry, r2_entry = _calculate_watershed_cn(ws_name, ws_df)
        result1_data.append(r1_entry)
        result2_data.append(r2_entry)

    logger.info(f"계산 완료: {len(watersheds)}개 소유역")
    return result1_data, result2_data, null_cn_rows


def calculate_grouped_results(layer: QgsVectorLayer, groups: dict) -> tuple:
    """
    유역합성 그룹별 CN값 계산.
    groups: {그룹명: [소유역1, 소유역2, ...]}
    반환: (grouped_result1_data, grouped_result2_data)
    """
    df = layer_to_dataframe(layer)
    grouped_result1 = []
    grouped_result2 = []

    for group_name, members in groups.items():
        group_df = df[df['소유역명'].isin(members)]
        if group_df.empty:
            logger.warning(f"유역합성 '{group_name}': 멤버 소유역 데이터 없음 ({members})")
            continue
        r1_entry, r2_entry = _calculate_watershed_cn(group_name, group_df)
        grouped_result1.append(r1_entry)
        grouped_result2.append(r2_entry)

    logger.info(f"유역합성 계산 완료: {len(grouped_result1)}개 그룹")
    return grouped_result1, grouped_result2


# ──────────────────────────────────────────────────────────────────────────────
# openpyxl 스타일 헬퍼
# ──────────────────────────────────────────────────────────────────────────────

def _make_styles():
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    font    = Font(name='맑은 고딕', size=11)
    align   = Alignment(horizontal='center', vertical='center')
    thin    = Side(style='thin')
    b_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_ltb   = Border(left=thin,             top=thin, bottom=thin)
    b_tb    = Border(             top=thin, bottom=thin)
    b_rtb   = Border(right=thin, top=thin, bottom=thin)
    fill_pink   = PatternFill(patternType='solid', fgColor='FFFF00FF')
    fill_cyan   = PatternFill(patternType='solid', fgColor='FFCCFFFF')
    fill_silver = PatternFill(patternType='solid', fgColor='FFC0C0C0')
    num_fmt = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
    return dict(
        font=font, align=align,
        b_all=b_all, b_ltb=b_ltb, b_tb=b_tb, b_rtb=b_rtb,
        fill_pink=fill_pink, fill_cyan=fill_cyan, fill_silver=fill_silver,
        num_fmt=num_fmt,
    )


def _sc(cell, value, font, align, border, fill=None, num_fmt=None):
    """스타일 적용 후 셀 값 설정."""
    cell.value  = value
    cell.font   = font
    cell.alignment = align
    cell.border = border
    if fill:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt


# ──────────────────────────────────────────────────────────────────────────────
# result1.xlsx 내보내기
# ──────────────────────────────────────────────────────────────────────────────

def export_result1(result1_data: list, path: str):
    """result1.xlsx: 소유역별 토지이용 상세 CN값 표 저장."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl이 필요합니다.")

    s = _make_styles()
    fn, al, nm = s['font'], s['align'], s['num_fmt']
    ball, bltb, brtb = s['b_all'], s['b_ltb'], s['b_rtb']
    fpink, fcyan = s['fill_pink'], s['fill_cyan']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # ── 1행: 전체 TYPE 헤더 ───────────────────────────────────────────────────
    r = 1
    _sc(ws.cell(r, 1), '소유역명', fn, al, ball, num_fmt=nm)
    _sc(ws.cell(r, 2), None,       fn, al, ball, num_fmt=nm)

    for idx, label in enumerate(['TYPE A', 'TYPE B', 'TYPE C', 'TYPE D']):
        c = 3 + idx * 2
        _sc(ws.cell(r, c),   label, fn, al, bltb, num_fmt=nm)
        _sc(ws.cell(r, c+1), None,  fn, al, brtb, num_fmt=nm)
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)

    for c in [11, 12, 13]:
        _sc(ws.cell(r, c), None, fn, al, ball, num_fmt=nm)

    # ── 소유역별 블록 ─────────────────────────────────────────────────────────
    r = 2
    for ws_data in result1_data:
        # 헤더 행 (소유역명 + 컬럼 레이블)
        _sc(ws.cell(r, 1), ws_data['watershed'], fn, al, ball, fill=fpink, num_fmt=nm)
        col_hdrs = ['토지이용','면적','CN','면적','CN','면적','CN','면적','CN','총면적','AMC2 CN','AMC3 CN']
        for i, h in enumerate(col_hdrs):
            _sc(ws.cell(r, i+2), h, fn, al, ball, fill=fcyan, num_fmt=nm)
        r += 1

        # 토지이용 데이터 행
        for row_data in ws_data['rows']:
            _sc(ws.cell(r, 1), None,               fn, al, ball, num_fmt=nm)
            _sc(ws.cell(r, 2), row_data['land_use'],fn, al, ball, num_fmt=nm)
            vals = [
                row_data['A_area'], row_data['A_cn'],
                row_data['B_area'], row_data['B_cn'],
                row_data['C_area'], row_data['C_cn'],
                row_data['D_area'], row_data['D_cn'],
                row_data['total_area'],
                row_data['amc2_cn'],
                row_data['amc3_cn'],
            ]
            for i, v in enumerate(vals):
                _sc(ws.cell(r, i+3), v, fn, al, ball, num_fmt=nm)
            r += 1

        # 합계 행
        _sc(ws.cell(r, 1), None, fn, al, ball, num_fmt=nm)
        _sc(ws.cell(r, 2), None, fn, al, ball, num_fmt=nm)
        summary_vals = [
            ws_data['total_A'] or None, None,
            ws_data['total_B'] or None, None,
            ws_data['total_C'] or None, None,
            ws_data['total_D'] or None, None,
            ws_data['total_area'],
            ws_data['amc2_cn'],
            ws_data['amc3_cn'],
        ]
        for i, v in enumerate(summary_vals):
            _sc(ws.cell(r, i+3), v, fn, al, ball, num_fmt=nm)
        r += 1

        # 구분 빈 행
        for c in range(1, 14):
            _sc(ws.cell(r, c), None, fn, al, ball, num_fmt=nm)
        r += 1

    ws.column_dimensions['L'].width = 9.375
    ws.column_dimensions['N'].width = 11.625

    wb.save(path)
    logger.info(f"result1.xlsx 저장: {path}")


# ──────────────────────────────────────────────────────────────────────────────
# result2.xlsx 내보내기
# ──────────────────────────────────────────────────────────────────────────────

def export_result2(result2_data: list, path: str):
    """result2.xlsx: 유역별 CN값 요약 저장."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl이 필요합니다.")

    s = _make_styles()
    fn, al, nm = s['font'], s['align'], s['num_fmt']
    ball, bltb, btb, brtb = s['b_all'], s['b_ltb'], s['b_tb'], s['b_rtb']
    fcyan, fsilver = s['fill_cyan'], s['fill_silver']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 1행: 타이틀 (A1:D1 병합)
    _sc(ws.cell(1, 1), ' 유역별 CN정리', fn, al, bltb, fill=fsilver, num_fmt=nm)
    _sc(ws.cell(1, 2), None,             fn, al, btb,  num_fmt=nm)
    _sc(ws.cell(1, 3), None,             fn, al, btb,  num_fmt=nm)
    _sc(ws.cell(1, 4), None,             fn, al, brtb, num_fmt=nm)
    ws.merge_cells('A1:D1')

    # 2행: 컬럼 헤더
    for i, h in enumerate(['유역', '총면적', 'AMC2 CN', 'AMC3 CN']):
        _sc(ws.cell(2, i+1), h, fn, al, ball, fill=fcyan, num_fmt=nm)

    # 데이터 행
    for ri, row in enumerate(result2_data):
        r = ri + 3
        _sc(ws.cell(r, 1), row['watershed'],  fn, al, ball, num_fmt=nm)
        _sc(ws.cell(r, 2), row['total_area'], fn, al, ball, num_fmt=nm)
        _sc(ws.cell(r, 3), row['amc2_cn'],    fn, al, ball, num_fmt=nm)
        _sc(ws.cell(r, 4), row['amc3_cn'],    fn, al, ball, num_fmt=nm)

    wb.save(path)
    logger.info(f"result2.xlsx 저장: {path}")


# ──────────────────────────────────────────────────────────────────────────────
# 통합 내보내기 (result1 + result2 → 단일 xlsx 파일, 2개 시트)
# ──────────────────────────────────────────────────────────────────────────────

def export_results(result1_data: list, result2_data: list, path: str,
                   *, grouped_result1: list = None, grouped_result2: list = None):
    """results.xlsx: result1 시트 + result2 시트를 하나의 파일에 저장."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl이 필요합니다.")

    s = _make_styles()
    fn, al, nm = s['font'], s['align'], s['num_fmt']
    ball, bltb, btb, brtb = s['b_all'], s['b_ltb'], s['b_tb'], s['b_rtb']
    fpink, fcyan, fsilver = s['fill_pink'], s['fill_cyan'], s['fill_silver']

    wb = Workbook()

    # ── Sheet: result1 ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'result1'

    r = 1
    _sc(ws1.cell(r, 1), '소유역명', fn, al, ball, num_fmt=nm)
    _sc(ws1.cell(r, 2), None,       fn, al, ball, num_fmt=nm)
    for idx, label in enumerate(['TYPE A', 'TYPE B', 'TYPE C', 'TYPE D']):
        c = 3 + idx * 2
        _sc(ws1.cell(r, c),   label, fn, al, bltb, num_fmt=nm)
        _sc(ws1.cell(r, c+1), None,  fn, al, brtb, num_fmt=nm)
        ws1.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
    for c in [11, 12, 13]:
        _sc(ws1.cell(r, c), None, fn, al, ball, num_fmt=nm)

    r = 2
    for ws_data in result1_data:
        _sc(ws1.cell(r, 1), ws_data['watershed'], fn, al, ball, fill=fpink, num_fmt=nm)
        col_hdrs = ['토지이용','면적','CN','면적','CN','면적','CN','면적','CN','총면적','AMC2 CN','AMC3 CN']
        for i, h in enumerate(col_hdrs):
            _sc(ws1.cell(r, i+2), h, fn, al, ball, fill=fcyan, num_fmt=nm)
        r += 1

        for row_data in ws_data['rows']:
            _sc(ws1.cell(r, 1), None,                fn, al, ball, num_fmt=nm)
            _sc(ws1.cell(r, 2), row_data['land_use'], fn, al, ball, num_fmt=nm)
            vals = [
                row_data['A_area'], row_data['A_cn'],
                row_data['B_area'], row_data['B_cn'],
                row_data['C_area'], row_data['C_cn'],
                row_data['D_area'], row_data['D_cn'],
                row_data['total_area'], row_data['amc2_cn'], row_data['amc3_cn'],
            ]
            for i, v in enumerate(vals):
                _sc(ws1.cell(r, i+3), v, fn, al, ball, num_fmt=nm)
            r += 1

        _sc(ws1.cell(r, 1), None, fn, al, ball, num_fmt=nm)
        _sc(ws1.cell(r, 2), None, fn, al, ball, num_fmt=nm)
        summary_vals = [
            ws_data['total_A'] or None, None,
            ws_data['total_B'] or None, None,
            ws_data['total_C'] or None, None,
            ws_data['total_D'] or None, None,
            ws_data['total_area'], ws_data['amc2_cn'], ws_data['amc3_cn'],
        ]
        for i, v in enumerate(summary_vals):
            _sc(ws1.cell(r, i+3), v, fn, al, ball, num_fmt=nm)
        r += 1

        for c in range(1, 14):
            _sc(ws1.cell(r, c), None, fn, al, ball, num_fmt=nm)
        r += 1

    # ── 유역합성 블록 (result1) ──────────────────────────────────────────────
    if grouped_result1:
        from openpyxl.styles import PatternFill
        fill_green = PatternFill(patternType='solid', fgColor='FFCCFFCC')

        # 구분 헤더
        _sc(ws1.cell(r, 1), '【유역합성】', fn, al, ball, fill=fsilver, num_fmt=nm)
        for c in range(2, 14):
            _sc(ws1.cell(r, c), None, fn, al, ball, fill=fsilver, num_fmt=nm)
        r += 1

        for ws_data in grouped_result1:
            _sc(ws1.cell(r, 1), ws_data['watershed'], fn, al, ball, fill=fill_green, num_fmt=nm)
            col_hdrs = ['토지이용','면적','CN','면적','CN','면적','CN','면적','CN','총면적','AMC2 CN','AMC3 CN']
            for i, h in enumerate(col_hdrs):
                _sc(ws1.cell(r, i+2), h, fn, al, ball, fill=fcyan, num_fmt=nm)
            r += 1

            for row_data in ws_data['rows']:
                _sc(ws1.cell(r, 1), None, fn, al, ball, num_fmt=nm)
                _sc(ws1.cell(r, 2), row_data['land_use'], fn, al, ball, num_fmt=nm)
                vals = [
                    row_data['A_area'], row_data['A_cn'],
                    row_data['B_area'], row_data['B_cn'],
                    row_data['C_area'], row_data['C_cn'],
                    row_data['D_area'], row_data['D_cn'],
                    row_data['total_area'], row_data['amc2_cn'], row_data['amc3_cn'],
                ]
                for i, v in enumerate(vals):
                    _sc(ws1.cell(r, i+3), v, fn, al, ball, num_fmt=nm)
                r += 1

            _sc(ws1.cell(r, 1), None, fn, al, ball, num_fmt=nm)
            _sc(ws1.cell(r, 2), None, fn, al, ball, num_fmt=nm)
            summary_vals = [
                ws_data['total_A'] or None, None,
                ws_data['total_B'] or None, None,
                ws_data['total_C'] or None, None,
                ws_data['total_D'] or None, None,
                ws_data['total_area'], ws_data['amc2_cn'], ws_data['amc3_cn'],
            ]
            for i, v in enumerate(summary_vals):
                _sc(ws1.cell(r, i+3), v, fn, al, ball, num_fmt=nm)
            r += 1

            for c in range(1, 14):
                _sc(ws1.cell(r, c), None, fn, al, ball, num_fmt=nm)
            r += 1

    ws1.column_dimensions['L'].width = 9.375
    ws1.column_dimensions['N'].width = 11.625

    # ── result2 블록 (result1 시트의 O열부터 작성) ───────────────────────────
    R2_COL = 15  # 'O'

    _sc(ws1.cell(1, R2_COL),   ' 유역별 CN정리', fn, al, bltb, fill=fsilver, num_fmt=nm)
    _sc(ws1.cell(1, R2_COL+1), None,             fn, al, btb,  num_fmt=nm)
    _sc(ws1.cell(1, R2_COL+2), None,             fn, al, btb,  num_fmt=nm)
    _sc(ws1.cell(1, R2_COL+3), None,             fn, al, brtb, num_fmt=nm)
    ws1.merge_cells(start_row=1, start_column=R2_COL, end_row=1, end_column=R2_COL+3)

    for i, h in enumerate(['유역', '총면적', 'AMC2 CN', 'AMC3 CN']):
        _sc(ws1.cell(2, R2_COL+i), h, fn, al, ball, fill=fcyan, num_fmt=nm)

    for ri, row in enumerate(result2_data):
        r2 = ri + 3
        _sc(ws1.cell(r2, R2_COL),   row['watershed'],  fn, al, ball, num_fmt=nm)
        _sc(ws1.cell(r2, R2_COL+1), row['total_area'], fn, al, ball, num_fmt=nm)
        _sc(ws1.cell(r2, R2_COL+2), row['amc2_cn'],    fn, al, ball, num_fmt=nm)
        _sc(ws1.cell(r2, R2_COL+3), row['amc3_cn'],    fn, al, ball, num_fmt=nm)

    # ── 유역합성 (result2) ───────────────────────────────────────────────────
    if grouped_result2:
        from openpyxl.styles import PatternFill
        fill_green = PatternFill(patternType='solid', fgColor='FFCCFFCC')

        r2_next = len(result2_data) + 3
        # 빈 행
        for c in range(R2_COL, R2_COL+4):
            _sc(ws1.cell(r2_next, c), None, fn, al, ball, num_fmt=nm)
        r2_next += 1
        # 구분 헤더
        _sc(ws1.cell(r2_next, R2_COL), '【유역합성】', fn, al, ball, fill=fsilver, num_fmt=nm)
        for c in range(R2_COL+1, R2_COL+4):
            _sc(ws1.cell(r2_next, c), None, fn, al, ball, fill=fsilver, num_fmt=nm)
        r2_next += 1

        for row in grouped_result2:
            _sc(ws1.cell(r2_next, R2_COL),   row['watershed'],  fn, al, ball, fill=fill_green, num_fmt=nm)
            _sc(ws1.cell(r2_next, R2_COL+1), row['total_area'], fn, al, ball, num_fmt=nm)
            _sc(ws1.cell(r2_next, R2_COL+2), row['amc2_cn'],    fn, al, ball, num_fmt=nm)
            _sc(ws1.cell(r2_next, R2_COL+3), row['amc3_cn'],    fn, al, ball, num_fmt=nm)
            r2_next += 1

    wb.save(path)
    logger.info(f"results.xlsx 저장: {path}")
